#!/usr/bin/env python
# coding: utf-8

# In[ ]:

''' PORT REGRESSION TESTER v2 '''
''' Author: MMAIOCCHI1 '''

import re
import os
import PIL
import csv
import time
import string
import openpyxl
import datetime
import win32gui
import warnings
import itertools
import numpy as np
import pandas as pd
from tkinter import *
from os import listdir
import win32com.client
import pyautogui as pag
import ipywidgets as widgets
from openpyxl import Workbook
from ipywidgets import Layout
from tkinter import messagebox
from os.path import isfile, join
from openpyxl.formatting import Rule
from datetime import datetime, timedelta
from stat import S_ISREG, ST_CTIME, ST_MODE
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Side, Font, Color, PatternFill


''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''
#set to 0 to see error outputs in the app
debugMode = 1

#files download folder of the bloomberg terminal
download_dir = "C:\\blp\\data\\"
#download_dir = "C:\\Users\\traveler\\AppData\\Local\\Temp\\Bloomberg\\data\\"


''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''

path = os.getcwd()
folder = path + "\\results"
daily_report = datetime.today().strftime("%m%d%y")
path_today = folder + "\\" + daily_report
path_report1 = path_today + "\\prod_reports"
path_report2 = path_today + "\\qa_reports"
path_results = path_today + "\\final_reports"
path_fails = path_today + "\\failures"

def create_folders():
    if not os.path.exists(folder):
        os.makedirs(folder)
    if not os.path.exists(path_today):
        os.makedirs(path_today)
    if not os.path.exists(path_report1):
        os.mkdir(path_report1)
    if not os.path.exists(path_report2):
        os.mkdir(path_report2)
    if not os.path.exists(path_results):
        os.mkdir(path_results)
    if not os.path.exists(path_fails):
        os.mkdir(path_fails)

def create_template():
    sheet = workbook.active
    recap_sheet = workbook['Sheet']
    recap_sheet.title = 'Results'
    recap_sheet.sheet_view.showGridLines = False
    recap_sheet['B2'].value = 'TESTS RESULTS'
    recap_sheet['B4'].value = 'Test Name'
    recap_sheet['C4'].value = 'Errors'
    recap_sheet.column_dimensions['B'].width = 40
    recap_sheet.column_dimensions['C'].width = 23
    for row in range(4,1000):
        recap_sheet.cell(row, 3).alignment = Alignment(horizontal='center', wrap_text=False)

            
            
''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''
            
class window_mgr():
    
    def __init__ (self):
        self._handle = None

    def find_window(self, class_name, window_name=None):
        self._handle = win32gui.FindWindow(class_name, window_name)

    def _window_enum_callback(self, hwnd, wildcard):
        if re.match(wildcard, str(win32gui.GetWindowText(hwnd))) is not None:
            self._handle = hwnd

    def find_window_wildcard(self, wildcard):
        self._handle = None
        win32gui.EnumWindows(self._window_enum_callback, wildcard)

    def set_foreground_k(self):   
        shell = win32com.client.Dispatch("WScript.Shell")
        shell.SendKeys('')
        win32gui.SetForegroundWindow(self._handle)
        time.sleep(t*1)
    
    def set_foreground(self):   
        win32gui.SetForegroundWindow(self._handle)
        
    def open_bbg_1(self):
        win.find_window_wildcard("1-BLOOMBERG")
        win.set_foreground_k()
        time.sleep(t*0.5)
        
    def open_bbg_2(self):
        win.find_window_wildcard("2-BLOOMBERG")
        win.set_foreground_k()
        time.sleep(t*0.5)
        
    def open_bbg_3(self):
        win.find_window_wildcard("3-BLOOMBERG")
        win.set_foreground_k()
        time.sleep(t*0.5)
        
    def open_bbg_4(self):
        win.find_window_wildcard("4-BLOOMBERG")
        win.set_foreground_k()
        time.sleep(t*0.5)
        
    def open_excel_file(self):
        win.find_window_wildcard(".*grid.*")
        win.set_foreground()
        time.sleep(t*1)
        
    def open_excel_formatted(self):
        win.find_window_wildcard(".* Excel.*")
        win.set_foreground()
        time.sleep(t*1)
            
win = window_mgr()

''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''
            
class bbg_mgr():

    
    def __init__(self,r=None,check=None,tmp=None,ptf=None,bmk=None,tab=None,
                 subt=None,view=None,day=None,ccy=None,bkdn=None,model=None,
                 unit=None,clvl=None,hz=None,breg=None,scen=None,m1=None,m2=None):
        
        self.r = r
        self.check = check
        self.tmp = tmp
        self.ptf = ptf
        self.bmk = bmk
        self.tab = tab
        self.subt = subt
        self.view = view
        self.day = day
        self.ccy = ccy
        self.bkdn = bkdn
        self.model = model
        self.unit = unit
        self.clvl = clvl
        self.hz = hz
        self.breg = breg
        self.scen = scen
        self.m1 = m1
        self.m2 = m2
        if day == 'None': self.day = 'Default'
    
    def press_go(self):
        pag.press('enter')
        pag.press('enter')
        time.sleep(t*0.5)
    
    def test_terminal(self):
        win.open_bbg_1()
        pag.write('THIS IS JUST A TEST. ')
        time.sleep(3)
        pag.write('CANCELING IN 5 SECONDS...')
        time.sleep(5)
        pag.press('esc')
        pag.press('esc')
        
    def test_excel(self):
        os.system('start excel.exe /e/')
        time.sleep(10)
        pag.hotkey('alt','fn', 'f4')
        time.sleep(2)
        
    def choose_BREG(self):
        if self.breg == 'wave1':
            breg = 'bbit_infield_use_dtl_for_equity_fields_wave_1'
        elif self.breg == 'wave2':
            breg = 'bbit_infield_use_dtl_for_equity_fields_wave_2'
        elif self.breg == 'wave3':
            breg = 'bbit_infield_use_dtl_for_equity_fields_wave_3'
        pag.write(breg)
            
    def setup_BREG(self):
        if self.breg == 'None':
            pass
        else:
            win.open_bbg_1()
            time.sleep(t*1)
            pag.write('9994')
            time.sleep(t*1)
            self.press_go()
            pag.write('5')
            time.sleep(t*1)
            self.press_go()
            time.sleep(t*1)
            pag.press('tab')
            time.sleep(t*1)
            self.choose_BREG()
            time.sleep(t*1)
            pag.press('tab')
            time.sleep(t*1)
            pag.write('TRUE')
            time.sleep(t*1)
            self.press_go()
            pag.press('1')
            self.press_go()
            time.sleep(t*1)
            pag.press('2')
            self.press_go()
            time.sleep(t*1)
            pag.press('1')
            self.press_go()
            time.sleep(t*1)
        
    def bbg_fnc(self, fnc):
        win.open_bbg_1()
        pag.write(fnc)
        self.press_go()
    
    def iter_tab(self, n):
        for _ in itertools.repeat(None, n):
            time.sleep(t*0.25)
            pag.press('tab')
    
    def select_ptf(self):
        pag.write(self.ptf)
        time.sleep(t*0.5)
        pag.press('f12')
        self.press_go()
        
    def select_ptf_prod(self):
        win.open_bbg_1()
        self.select_ptf()
        
    def select_ptf_dtl(self):
        win.open_bbg_1()
        self.select_ptf()
    
    def select_ptf_qa(self):
        win.open_bbg_2()
        self.select_ptf()
    
    def open_PORT_prod(self):
        win.open_bbg_1()
        time.sleep(t*0.5)
        pag.write("PORT " + self.tab + " V " + self.view + " /QA")
        self.press_go()
        
    def open_PORT_dtl(self):
        win.open_bbg_1()
        time.sleep(t*0.5)
        pag.write("RRRR PORT " + self.m2 + " " + self.tab + " V " + self.view + " /QA")
        self.press_go()
        time.sleep(t*6)
    
    def open_PORT_qa(self):
        win.open_bbg_2()
        time.sleep(t*0.5)
        pag.write("RRRR PORT " + self.m2 + " " + self.tab + " V " + self.view + " /QA")
        self.press_go()
        time.sleep(t*6)
        
    def change_subtab(self):
        if self.tab == 'HP':
            if self.subt == 'MainView':
                pag.write('30')
            elif self.subt == 'TotalReturn':
                pag.write('31')
            elif self.subt == 'PeriodAnalysis':
                pag.write('32')
            elif self.subt == 'SeasonalAnalysis':
                pag.write('33')
            elif self.subt == 'attributioStatisticalSummary':  
                pag.write('34')
        else:
            if self.subt == 'MainView':
                pag.write('30')
            elif self.subt == 'Summary':
                pag.write('31')
            elif self.subt == 'Factors':
                pag.write('32')
            elif self.subt == 'RiskBets':
                pag.write('33')
            elif self.subt == 'Trends':  
                pag.write('34')
            elif self.subt == 'Exposures':
                pag.write('35')
        time.sleep(t*0.5)
        self.press_go()
        time.sleep(t*2)
        
    def setup_all_widgets(self, machine):
        time.sleep(t*2)
        dropdown_values_list = [
                                    self.bmk,
                                    self.bkdn,
                                    self.ccy,
                                    self.day,
                                    self.model,
                                    self.unit,
                                    self.hz,
                                    self.clvl,
                                    self.scen
                                ]

        if self.tab in ['HD','HP', 'CH']:
            final_values_list = dropdown_values_list[0:4]
        if self.tab == 'PA':
            final_values_list = dropdown_values_list[0:5]
        if self.tab == 'TE':
            final_values_list = dropdown_values_list[0:7]
        if self.tab == 'VR':
            final_values_list = dropdown_values_list[0:6] + [dropdown_values_list[7]] + [dropdown_values_list[6]]
        if self.tab == 'SA':
            final_values_list = dropdown_values_list[0:5] + [dropdown_values_list[8]]

        if final_values_list.count('Default') - len(final_values_list) != 0:
            self.iter_tab(2)
            for i in range(0, len(final_values_list)):

                    '''handle exceptions'''
                    #handle start date in PA tab
                    if self.tab == 'PA' and i == 3:
                        self.iter_tab(4)
                    #handle day dropdown when prev_close
                    if i == 3 and final_values_list[i] == 'Default':
                        self.iter_tab(2)
                    #handle V8 dropdown until it's confirmed
                    if i == 4 and self.tab in ['TE','VR','SA']:
                        if self.tab == 'TE':
                            self.iter_tab(1)
                        elif self.tab != 'TE' and machine == 'QA':
                            self.iter_tab(1)
                        else:
                            pass
                    '''end of exceptions'''

                    #IF VALUE IS NOT DEFAULT
                    if final_values_list[i] != 'Default':
                        #handle day typing
                        if i == 3:
                            mm = str(self.day[5:7])
                            dd = str(self.day[8:10])
                            yy = str(self.day[2:4])
                            pag.write(mm)
                            time.sleep(0.5)
                            pag.write(dd)
                            time.sleep(0.5)
                            pag.write(yy)
                            time.sleep(0.5)
                            if self.tab != 'TE':
                                self.iter_tab(1)
                        else:
                            if final_values_list[i] == final_values_list[-1]:
                                pag.write(final_values_list[i])
                            else:
                                pag.write(final_values_list[i])
                                self.iter_tab(1)
                    #IF VALUE IS DEFAULT    
                    else:
                        self.iter_tab(1)

    def setup_widgets_prod(self):
        win.open_bbg_1()
        self.change_subtab()
        win.open_bbg_1()
        self.setup_all_widgets('PROD')
        self.press_go()
        time.sleep(t*2)
        
    def setup_widgets_qa(self):
        win.open_bbg_2()
        self.change_subtab()
        win.open_bbg_2()
        self.setup_all_widgets('QA')
        self.press_go()
        time.sleep(t*2)
        
    def setup_widgets_custom(self, formulas):
        if formulas == 'withBREG':
            win.open_bbg_1()
            self.change_subtab()
            win.open_bbg_1()
            self.setup_all_widgets('QA')
            self.press_go()
            time.sleep(t*2)
        else:
            win.open_bbg_2()
            self.change_subtab()
            win.open_bbg_2()
            self.setup_all_widgets('QA')
            self.press_go()
            time.sleep(t*2)
    
    def export(self, machine):
        if machine == 'PROD':
            win.open_bbg_1()
        else:
            win.open_bbg_2()
        if self.tmp == 'Current Tab (Unformatted xls)':
            time.sleep(t*2)
            pag.write('12')
            self.press_go()
            pag.write('6')
            self.press_go()
            pag.write('1')
            self.press_go()
        else:
            time.sleep(t*2)
            pag.write('12')
            self.press_go()
            pag.write('6')
            self.press_go()
            pag.write('2')
            self.press_go()
        time.sleep(t*5)
        
    def closeExcel(self):
        if self.tmp == 'Current Tab (Unformatted xls)': 
            win.open_excel_file()
            time.sleep(1)
            pag.hotkey('ctrl','fn', 'f4')
            try:
                win.open_excel_file()
                time.sleep(1)
                pag.hotkey('ctrl','fn', 'f4')
            except:
                pass
        else:
            win.open_excel_formatted()
            time.sleep(1)
            pag.hotkey('alt','fn', 'f4')
            try:
                win.open_excel_formatted()
                time.sleep(1)
                pag.hotkey('alt','fn', 'f4')   
            except:
                pass
        time.sleep(2)
                          
    def export_mainview(self, machine):
        while True:
            try:
                x = 0
                while x < 10:
                    try:
                        x = x + 1
                        time.sleep(t*2)
                        win.open_excel_file()
                        break
                    except:
                        continue
                #if generated, load it to see inside
                file = win32gui.GetWindowText(win32gui.GetForegroundWindow()).split()[0]
                os.chdir(download_dir)
                wb = openpyxl.load_workbook(file + '.xlsx')
                sheet = wb['Worksheet']
                #if empty export again for maximum 10 times
                y = 0
                while str(str(sheet['B2'].value) + str(sheet['B3'].value) + str(sheet['B4'].value) +
                          str(sheet['B5'].value) + str(sheet['B6'].value) + str(sheet['C2'].value) +
                          str(sheet['C3'].value) + str(sheet['C4'].value) + str(sheet['C5'].value) +
                          str(sheet['C6'].value) + str(sheet['D2'].value) + str(sheet['D3'].value) +
                          str(sheet['D4'].value) + str(sheet['D5'].value) + str(sheet['D6'].value) +
                          str(sheet['E2'].value)) == 'NoneNoneNoneNoneNoneNoneNoneNoneNoneNoneNoneNoneNoneNoneNoneNone' and y < 20 :
                    y = y + 1
                    self.closeExcel()
                    self.export(machine)
                    x = 0
                    while x < 5:
                        try:
                            x = x + 1
                            time.sleep(t*5)
                            win.open_excel_file()
                            break
                        except:
                            continue
                    #if generated, load it to see inside
                    file = win32gui.GetWindowText(win32gui.GetForegroundWindow()).split()[0]
                    os.chdir(download_dir)
                    wb = openpyxl.load_workbook(file + '.xlsx')
                    sheet = wb['Worksheet']
                #kicks an error if tried to export not-empty report more than 10 times
                if y == 10:
                    self.closeExcel()
                else:
                    pass
                break
            except:
                break
    
    def export_else(self, machine):
        export_iter = 0
        while export_iter < 5:
            export_iter = export_iter +  1
            x = 0
            while x < 5:
                x = x + 1
                try: 
                    time.sleep(t*5)
                    win.open_excel_file()
                    break
                except:
                    continue
            file_check = win32gui.GetWindowText(win32gui.GetForegroundWindow()).split()[0][0:4]
            if file_check == 'grid':
                file = win32gui.GetWindowText(win32gui.GetForegroundWindow()).split()[0]
                os.chdir(download_dir)
                wb = openpyxl.load_workbook(file + '.xlsx')
                sheet = wb['Worksheet']
                y = 0
                #if empty export again for maximum 10 times
                while str(str(sheet['B2'].value) + str(sheet['B3'].value) + str(sheet['B4'].value) +
                          str(sheet['B5'].value) + str(sheet['B6'].value) + str(sheet['C2'].value) +
                          str(sheet['C3'].value) + str(sheet['C4'].value) + str(sheet['C5'].value) +
                          str(sheet['C6'].value) + str(sheet['D2'].value) + str(sheet['D3'].value) +
                          str(sheet['D4'].value)) == 'NoneNoneNoneNoneNoneNoneNoneNoneNoneNoneNoneNoneNone' and y < 10 :
                    y = y + 1
                    self.closeExcel()
                    self.export(machine)
                    z = 0
                    while z < 5:
                        z = z + 1
                        try:
                            time.sleep(t*5)
                            win.open_excel_file()
                            break
                        except:
                            continue
                    file = win32gui.GetWindowText(win32gui.GetForegroundWindow()).split()[0]
                    os.chdir(download_dir)
                    wb = openpyxl.load_workbook(file + '.xlsx')
                    sheet = wb['Worksheet']
                #kicks an error if tried to export not-empty report more than 5 times
                if y == 10:
                    self.closeExcel()
                else:
                    pass
                break
            else:
                self.export(machine)
                time.sleep(t*5)
                continue
                           
    def export_formatted(self):
        x = 0
        while x < 100:
            x = x + 1
            try:
                time.sleep(t*5)
                win.open_excel_formatted()
                break
            except:
                continue
                                  
    def export_loop(self, machine):
        #this contains all loops to export PORT reports
        if self.tmp == 'Current Tab (Unformatted xls)':
            if self.subt == 'MainView':
                self.export_mainview(machine)
            else:
                self.export_else(machine)                  
        else:
            self.export_formatted()
            
    def export_prod(self):
        self.export('PROD')
        self.export_loop('PROD')
        
    def export_qa(self):
        self.export('QA')
        self.export_loop('QA')
        
    def save_XLS(self, _id, window):
        if self.tmp == 'Current Tab (Unformatted xls)': 
            time.sleep(t*1)
            file = win32gui.GetWindowText(win32gui.GetForegroundWindow()).split()[0]
            os.chdir(download_dir)
            wb = openpyxl.load_workbook(file + '.xlsx')
            if window == 1:
                filename = str(path_report1 + '\\1_' + self.tab + '_' + self.subt + '_' + self.ptf + '_' + str(_id) +'.xlsx')
            else:
                filename = str(path_report2 + '\\2_' + self.tab + '_' + self.subt + '_' + self.ptf + '_' + str(_id) +'.xlsx')
            wb.save(filename = filename)
        else:
            time.sleep(t*5)
            file = win32gui.GetWindowText(win32gui.GetForegroundWindow())[0:8]
            fname = download_dir + file + ".xls"
            excel = win32com.client.gencache.EnsureDispatch('Excel.Application')
            wb = excel.Workbooks.Open(fname)
            if window == 1:
                wb.SaveAs(path_report1 + '\\1_' + self.tab + '_' + self.subt + '_' + self.ptf + '_' + str(_id) +'.xlsx',FileFormat = 51)
            else:
                wb.SaveAs(path_report2 + '\\2_' + self.tab + '_' + self.subt + '_' + self.ptf + '_' + str(_id) +'.xlsx',FileFormat = 51)
                 
    def df_comparison(self, r):
        if self.tmp == 'Current Tab (Unformatted xls)':
            df1 = pd.read_excel(path_report1 + '\\1_' + self.tab + '_' + self.subt + '_' + self.ptf + '_' + str(r) + '.xlsx',
                               header = None).fillna(0)
            df2 = pd.read_excel(path_report2 + '\\2_' + self.tab + '_' + self.subt + '_' + self.ptf + '_' + str(r) + '.xlsx',
                               header = None).fillna(0)
        else:
            df1 = pd.read_excel(path_report1 + '\\1_' + self.tab + '_' + self.subt + '_' + self.ptf + '_' + str(r) + '.xlsx',
                                header = None).fillna(0)[:-1]
            df2 = pd.read_excel(path_report2 + '\\2_' + self.tab + '_' + self.subt + '_' + self.ptf + '_' + str(r) + '.xlsx',
                                header = None).fillna(0)
            df2 = df2.iloc[:,0:len(df2.columns)+(df1.shape[1]-df2.shape[1])] #fix to avoid styling issues in QA
        #round to 6 decimals if data point in df is float
        df1 = df1.applymap(lambda x: round(x,6) if isinstance(x,float) else x)
        df2 = df2.applymap(lambda x: round(x,6) if isinstance(x,float) else x)
        #make comparison df
        df_comparison = (df1 == df2)
        for c in range(0,len(df_comparison.columns)):
            for rw in range(0,len(df_comparison.index)):
                if df_comparison.iloc[rw,c] == True:
                    df_comparison.iloc[rw,c] = df1.iloc[rw,c]
                else:
                    df_comparison.iloc[rw,c] = str(round(df1.iloc[rw,c],6)) + " <> " + str(round(df2.iloc[rw,c],6))
        return df_comparison
           
    def dump_results_to_excel(self, r):
        df_comparison = self.df_comparison(r)
        sheet_new = workbook.create_sheet(str(r) + '_' + self.ptf)
        sheet_new.sheet_view.showGridLines = False
        rows = dataframe_to_rows(df_comparison)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                 sheet_new.cell(row=r_idx, column=c_idx, value=value)
        #add formatting to new comparison sheet
        red_text = Font(color="9C0006")
        red_fill = PatternFill(bgColor="FFC7CE")
        dxf = DifferentialStyle(font=red_text, fill=red_fill)
        rule = Rule(type="containsText", operator="containsText", text="<>", dxf=dxf)
        rule.formula = ['NOT(ISERROR(SEARCH("<>",A1)))']
        sheet_new.conditional_formatting.add('A1:ZZ5000', rule)
        for row in sheet_new['A1:ZZ2']:
            for cell in row:
                cell.value = None
        for row in sheet_new['A1:A10000']:
            for cell in row:
                cell.value = None
        #recap results in summary page
        workbook['Results']['B' + str(r+5)].value = str(r) + '_' + self.check + '_' + self.ptf + '_' + self.tab + '_' + self.subt
        tot_err = []
        for c in range(0,len(df_comparison.columns)):
            errbycol = df_comparison[c].str.count("<>").sum()
            tot_err.append(errbycol)
        tot_err = sum(tot_err)
        workbook['Results']['C' + str(r+5)].value = round(tot_err,2)
        
    def summarize_byEqtFields(self, r):
        df_comparison = self.df_comparison(r)
        fields_list = pd.DataFrame(df_comparison.iloc[0].to_list(), columns = ['Equity Fields'])
        tot_err = []
        for c in range(0,len(df_comparison.columns)):
            errbycol = df_comparison[c].str.count("<>").sum()
            tot_err.append(errbycol)
        fields_list['Errors'] = tot_err
        #dump summary byEqtField to new sheet
        sheet_new = workbook.create_sheet(str(r) + '_' + self.ptf)
        sheet_new.sheet_view.showGridLines = False
        sheet_new.column_dimensions['B'].width = 20
        rows = dataframe_to_rows(fields_list)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                 sheet_new.cell(row=r_idx, column=c_idx, value=value)
        #recap results in summary page (only rows with errors)
        workbook['Results']['B' + str(r+5)].value = str(r) + '_' + self.check + '_' + self.ptf + '_' + self.tab + '_' + self.subt
        tot_err = sum(tot_err)
        workbook['Results']['C' + str(r+5)].value = round(tot_err,2)
        #add filtered comparison next to recap
        df_filtered = df_comparison[:1]
        for row in range(0, len(df_comparison.index)):
            some_list = df_comparison.iloc[row].to_list()
            some_list = list(map(str, some_list))
            if '<>' in '\t'.join(some_list):
                df_filtered = df_filtered.append(df_comparison[row:row+1])
        rows = dataframe_to_rows(df_filtered)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 5):
                 sheet_new.cell(row=r_idx, column=c_idx, value=value)
        #add formatting to new comparison sheet
        red_text = Font(color="9C0006")
        red_fill = PatternFill(bgColor="FFC7CE")
        dxf = DifferentialStyle(font=red_text, fill=red_fill)
        rule = Rule(type="containsText", operator="containsText", text="<>", dxf=dxf)
        rule.formula = ['NOT(ISERROR(SEARCH("<>",A1)))']
        sheet_new.conditional_formatting.add('A1:ZZ5000', rule)
        

    def PROD_vs_QA(self, r):
        self.select_ptf_prod()
        self.open_PORT_prod()
        self.select_ptf_qa()
        self.open_PORT_qa()
        self.setup_widgets_prod()
        self.setup_widgets_qa()
        self.export_prod()
        self.save_XLS(r,1)
        self.closeExcel()
        self.export_qa()
        self.save_XLS(r,2)
        self.closeExcel()
        self.df_comparison(r)
        self.dump_results_to_excel(r)
  
    def BREG_vs_noBREG(self, r):
        if self.breg != 'Default':
            self.select_ptf_dtl()
            self.open_PORT_dtl()
            self.setup_BREG()
            self.select_ptf_qa()
            self.open_PORT_qa()
        self.setup_widgets_custom('withBREG')
        self.setup_widgets_custom('noBREG')
        self.export_prod()
        self.save_XLS(r,1)
        self.closeExcel()
        self.export_qa()
        self.save_XLS(r,2)
        self.closeExcel()
        self.summarize_byEqtFields(r)
    
    #add here a new customized test
    def TEST(self,r):
        if self.check == 'PROD_vs_QA':
            self.PROD_vs_QA(r)
        elif self.check == 'BREG_vs_noBREG':
            self.BREG_vs_noBREG(r)
            
            
    ''''''''''''''''''''    
    '''Error handlers'''
    
    ''''''''''''''''''''  
    def err_handler(self, r):
        sheet_err = workbook.create_sheet(str(r) + '_' + self.ptf)
        sheet_err.sheet_view.showGridLines = False
        sheet_err['B2'].value = 'Unable to compare : reports have different number of rows/columns!'
        sheet_err['B3'].value = 'Check out the raw reports in the output folder'
        workbook['Results']['B' + str(r+5)].value = str(r) + '_' + self.check + '_' + self.ptf + '_' + self.tab + '_' + self.subt
        workbook['Results']['C' + str(r+5)].value = 'Error: reports with different # of rows/columns'
    
    def err_handler_manual(self, r):
        pag.FAILSAFE = False
        sheet_err = workbook.create_sheet(str(r) + '_' + self.ptf)
        sheet_err.sheet_view.showGridLines = False
        sheet_err['B2'].value = 'Interrupted manually by moving the mouse to the corner of the screen'
        workbook['Results']['B' + str(r+5)].value = str(r) + '_' + self.check + '_' + self.ptf + '_' + self.tab + '_' + self.subt
        workbook['Results']['C' + str(r+5)].value = 'Interrupted manually'
        pag.FAILSAFE = True
    
    def iteration_err_handler(self, r):
        time.sleep(1)
        myScreenshot = pag.screenshot()
        myScreenshot.save(path_fails + '\\' + str(r) + '_' + self.ptf + '.png')
        time.sleep(1)
        win.open_bbg_1()
        pag.press('esc')
        pag.press('esc')
        pag.press('esc')
        time.sleep(1)
        win.open_bbg_2()
        pag.press('esc')
        pag.press('esc')
        pag.press('esc')
        try:
            time.sleep(1)
            win.find_window_wildcard(".*Excel.*")
            win.set_foreground()
            bbg_rt.closeExcel()
        except:
            pass
        sheet_err = workbook.create_sheet(str(r) + '_' + self.ptf)
        sheet_err.sheet_view.showGridLines = False
        sheet_err['B2'].value = 'Something went wrong, please check screenshot ' + str(r) +' in FAILURES folder'
        workbook['Results']['B' + str(r+5)].value = str(r) + '_' + self.check + '_' + self.ptf + '_' + self.tab + '_' + self.subt
        workbook['Results']['C' + str(r+5)].value = 'Something went wrong'


''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''
        
class UI:
           
    tab_list = ['CH','HD','HP','PA','TE','VR','SA']
    ptf_list = ['H160819-6','H160819-12','H160819-52','U4153597-165','...']
    view_list = ['NX_EQUITY','NX_FIXINC','NX_BALANCED']
    subtab_list = ['MainView',' ','--TE tabs--','Summary', 'Factors', 'RiskBets', 'Trends', 'Exposure',
                  ' ','--HP tabs--','TotalReturn','PeriodAnalysis','SeasonalAnalysis','StatisticalSummary']
    test_type = ['PROD_vs_QA','BREG_vs_noBREG']
    qa_machine = ['2973','2974','2977']
    bmk_list = ['Default', 'None', 'MXWO', 'INDU']
    bkdn_list = ['Default', 'None', 'Market Cap Ranges', 'Security Type']
    ccy_list = ['Default', 'USD', 'EUR', 'GBP', 'JPY', 'CHF', 'CAD', 'HKD']
    day_list = ['Default','12/31/19','01/16/20','01/31/20','...']
    model_list = ['Default', ' ','----Risk Models----', 'Asian', 'Australian', 'Canadian',
                  'Chinese', 'Emerging','European', 'Global', 'Japanese', 'Latin America',
                  'US', 'Fixed Income','Bloomberg Risk Model (Global)', 'Bloomberg Risk Model (Regional)',
                  ' ','----Attribution Models----', 'Total Ret', 'Spread Ret','Excess Ret', 'Factor']
    unit_list = ['Default', 'P&L', 'Returns', 'Basis Points']
    clvl_list = ['Default', '90', '95', '99']
    hz_list = ['Default', '1 D', '1 W', '2 W', '1 M', '1 Q','1 Y']
    tmp_list = ['Current Tab (Unformatted xls)', 'Current Tab (xls)']
    scen_list = ['Default','All Scenarios', 'Equity Markets', 'Greece', 'Libya', 'Russian', 'Japan','Lehman']
    breg_list = ['None','wave1','wave2','wave3']

    
    tests_type = widgets.Dropdown(
        options=test_type,
        description='Test : ',
        disabled=False)

    tabs_list = widgets.Dropdown(
        options=tab_list,
        description='Function : ',
        disabled=False)

    subtabs_list = widgets.Dropdown(
        options=subtab_list,
        description='Tab : ',
        disabled=False)

    ptfs_list = widgets.Dropdown(
        options=ptf_list,
        description='PTF #: ',
        disabled=False)

    views_list = widgets.Dropdown(
        options = view_list,
        description='Views : ',
        disabled=False)   

    qa_machines = widgets.Dropdown(
        options=qa_machine,
        description='QA mach : ',
        disabled=False)

    bmks_list = widgets.Dropdown(
        options=bmk_list,
        description='Benchmark : ',
        disabled=False,
        value = 'Default')

    bkdns_list = widgets.Dropdown(
        options=bkdn_list,
        description='Breakdown : ',
        disabled=False,
        value = 'Default')

    ccys_list = widgets.Dropdown(
        options=ccy_list,
        description='Currency : ',
        disabled=False,
        value = 'Default')

    days_list = widgets.DatePicker(
        description='Day : ',
        disabled=False)
    
    models_list = widgets.Dropdown(
        options=model_list,
        description='Model : ',
        disabled=False,
        value = 'Default')

    units_list = widgets.Dropdown(
        options=unit_list,
        description='Unit : ',
        disabled=False,
        value = 'Default')

    clvls_list = widgets.Dropdown(
        options=clvl_list,
        description='Conf. Level : ',
        disabled=False,
        value = 'Default')

    hzs_list = widgets.Dropdown(
        options=hz_list,
        description='Horizon : ',
        disabled=False,
        value = 'Default')

    tmps_list = widgets.Dropdown(
        options=tmp_list,
        description='Temp : ',
        disabled=False)

    scens_list = widgets.Dropdown(
        options=scen_list,
        description='Scenario : ',
        disabled=False)
    
    bregs_list = widgets.Dropdown(
        options=breg_list,
        description='Bregs : ',
        disabled=False)

    start_button = widgets.Button(
        description='START REGRESSION TEST 🛠',
        disabled=False,
        button_style='danger',
        layout=widgets.Layout(width='80%'))

    ctrlfile_button = widgets.Button(
        description='OPEN CTRL FILE',
        disabled=False)
    
    turnoff_button = widgets.Checkbox(
        value=False,
        description='Turn off pc after long run')
    
    hidden_button_small = widgets.Button(
        description='hidden',
        layout=widgets.Layout(width='20%'))
    hidden_button_small.layout.visibility = 'hidden'
    
    hidden_button_large = widgets.Button(
        description='hidden',
        layout=widgets.Layout(width='50%'))
    hidden_button_large.layout.visibility = 'hidden'
    hidden_box = widgets.HBox([hidden_button_large])
    
    #widgets shown in tab 1
    ctrlfileTab = widgets.HBox([hidden_button_small, ctrlfile_button, hidden_button_small, turnoff_button])
    
    #widgets shown in tab 2
    box1 = widgets.HBox([ptfs_list, tabs_list, views_list])
    box2 = widgets.HBox([subtabs_list, bmks_list, bkdns_list])
    box3 = widgets.HBox([ccys_list, days_list, units_list])
    box4 = widgets.HBox([models_list, clvls_list, hzs_list ])
    box5 = widgets.HBox([scens_list, qa_machines, tmps_list])
    box6 = widgets.HBox([tests_type, bregs_list])
    widgetsTab = widgets.VBox([box6,hidden_box,box1,box2,box3,box4,box5])
   
    #start button
    box_start = widgets.HBox([hidden_button_large,start_button,hidden_button_large])
    
    
    
     
''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''
    
class worker():

    def open_ctrl_file(self):
        os.chdir(path)
        os.system('start excel.exe Control_file.xlsx') 
        
    def launch_regtest(self):
        global workbook
        workbook = Workbook()
        create_folders()
        create_template()
        final_file = path_results + '\\final_report_' + str(datetime.now().strftime("%H%M")) + '.xlsx'
        if inputs == 'CTRL FILE': #inputs is a global variable that changes by switching tab in the app UI
            xls = pd.ExcelFile(path + "\\Control_file.xlsx")
            df_custom = pd.read_excel(xls, 'CUSTOM')
        else:
            df_dict = {
                              "Test": UI.tests_type.value,
                              "Temp": UI.tmps_list.value,
                              "Name": '',
                              "Portfolio": UI.ptfs_list.value,
                              "Benchmark": UI.bmks_list.value,
                              "Tab": UI.tabs_list.value,
                              "Subtab": UI.subtabs_list.value,
                              "View": UI.views_list.value,
                              "As of": UI.days_list.value,
                              "SDA": '',
                              "Ccy": UI.ccys_list.value,
                              "Bkdn": UI.bkdns_list.value,
                              "Model": UI.models_list.value,
                              "Unit": UI.units_list.value,
                              "Clvl": UI.clvls_list.value,
                              "Hz": UI.hzs_list.value,
                              "Set": '',
                              "Breg": UI.bregs_list.value,
                              "Scen": UI.scens_list.value,
                              "Mach1": '',
                              "Mach2": UI.qa_machines.value
                       }
            
            df_dict = pd.DataFrame(df_dict.items()).T
            df_dict.columns = df_dict.iloc[0]
            df_custom = df_dict.drop(0)
            
        #run TEST for each row in the inputs dataframe
        for r in range(0,len(df_custom.index)):

            launch = bbg_mgr(
                                      r,
                                      df_custom.iloc[r,0],
                                      df_custom.iloc[r,1],
                                      df_custom.iloc[r,3],
                                      df_custom.iloc[r,4],
                                      df_custom.iloc[r,5],
                                      df_custom.iloc[r,6],
                                      df_custom.iloc[r,7],
                                  str(df_custom.iloc[r,8]),
                                      df_custom.iloc[r,10],
                                      df_custom.iloc[r,11],
                                      df_custom.iloc[r,12],
                                      df_custom.iloc[r,13],
                                  str(df_custom.iloc[r,14]),
                                      df_custom.iloc[r,15],
                                      df_custom.iloc[r,17],
                                      df_custom.iloc[r,18],
                                  str(df_custom.iloc[r,19]),
                                  str(df_custom.iloc[r,20])
                             )
            
            if debugMode == 0:
                launch.TEST(r)
            else:
                try:
                     launch.TEST(r) #MAIN FUNCTION
                except Exception as e:
                    if str(e) == 'Can only compare identically-labeled DataFrame objects':
                        launch.err_handler(r)
                    elif str(e).split(' ')[0] == 'PyAutoGUI':
                        launch.err_handler_manual(r)
                        break
                    elif str(e) == 'Can only use .str accessor with string values!':
                        pass
                    else:
                        launch.iteration_err_handler(r)


            #save result template after each iteration
            workbook.save(final_file)

        if UI.turnoff_button.value == True:
            workbook.save(final_file)
            os.system("shutdown /s /t 1")
        else:
            workbook.save(final_file)
            os.startfile(final_file)


''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''
    
class app():

    def run():

        #Switch tab to select different input types (ctrl file / widgets)
        tabs = widgets.Tab()
        tabs.children = [UI.ctrlfileTab,UI.widgetsTab]
        tabs.set_title(0, 'CTRL FILE')
        tabs.set_title(1, 'WIDGETS')

        def switch_inputs(*args):
            global inputs
            if tabs.selected_index == 0:
                inputs = 'CTRL FILE'
            else:
                inputs = 'WIDGETS'
        tabs.observe(switch_inputs)
        switch_inputs()
        
        #Control regtest speed using a slider
        speed_slider = widgets.FloatSlider(
                        value=1,
                        min=0.8,
                        max=1.5,
                        step=0.1,
                        description='Speed:',
                        disabled=False,
                        continuous_update=False,
                        orientation='horizontal',
                        readout=True,
                        readout_format='.1f',
                        )
        
        def switch_speed(*args):
            global t
            t = speed_slider.value
        speed_slider.observe(switch_speed)
        switch_speed()
        
        
        #Defines final app UI printed in the jupyter notebook
        finalUI = widgets.VBox([
                                tabs,
                                UI.hidden_box,
                                speed_slider,
                                UI.hidden_box,
                                UI.box_start
                               ])
        
        #launch worker class functions when pressing UI buttons
        UI.ctrlfile_button.on_click(worker.open_ctrl_file)
        UI.start_button.on_click(worker.launch_regtest)
        
        return finalUI
        
        
        
       


        
